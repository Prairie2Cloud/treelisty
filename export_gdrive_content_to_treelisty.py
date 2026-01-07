"""
TreeListy Google Drive Content Extractor - Knowledge Base Pattern
Downloads and extracts text content from Google Drive files for RAG.

Unlike export_google_drive_to_treelisty.py (metadata only), this script:
- Downloads actual file content
- Extracts text from Google Docs, PDFs, Word, Excel, etc.
- Creates a knowledge-base pattern tree with chunks
- Adds external IDs for Dashboard Trees merge
- Includes RAG metadata for semantic search

Setup:
1. Install: pip install google-auth google-auth-oauthlib google-auth-httplib2 google-api-python-client PyMuPDF python-docx openpyxl
2. Enable API: https://console.cloud.google.com/apis/library/drive.googleapis.com
3. Create credentials: https://console.cloud.google.com/apis/credentials (Desktop app OAuth 2.0)
4. Download credentials.json to this folder
5. Run: python export_gdrive_content_to_treelisty.py [folder_id] [--max-depth N] [--chunk-size N]

First run opens browser for authentication. Token saved for future runs.
"""

import os
import sys
import json
import re
import io
import argparse
from datetime import datetime
from typing import Optional, List, Dict, Any

# Fix Windows console encoding for emojis
if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8')

from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload

# Optional imports for text extraction
try:
    import fitz  # PyMuPDF
    HAS_PYMUPDF = True
except ImportError:
    HAS_PYMUPDF = False
    print("⚠️  PyMuPDF not installed. PDF extraction disabled. Install: pip install PyMuPDF")

try:
    from docx import Document as DocxDocument
    HAS_DOCX = True
except ImportError:
    HAS_DOCX = False
    print("⚠️  python-docx not installed. Word extraction disabled. Install: pip install python-docx")

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("⚠️  openpyxl not installed. Excel extraction disabled. Install: pip install openpyxl")

# Google Drive API scope (read-only)
SCOPES = ['https://www.googleapis.com/auth/drive.readonly']
TOKEN_FILE = 'token-drive.json'  # Separate from Gmail token

# Default chunk size in characters (roughly 250-500 tokens)
DEFAULT_CHUNK_SIZE = 1500
MAX_CHUNK_SIZE = 4000
MIN_CHUNK_SIZE = 200

# File types we can extract text from
EXTRACTABLE_TYPES = {
    # Google Workspace (export as text)
    'application/vnd.google-apps.document': {'export': 'text/plain', 'icon': '📘', 'name': 'Google Doc'},
    'application/vnd.google-apps.spreadsheet': {'export': 'text/csv', 'icon': '📗', 'name': 'Google Sheet'},
    'application/vnd.google-apps.presentation': {'export': 'text/plain', 'icon': '📙', 'name': 'Google Slides'},

    # Standard files (download and parse)
    'application/pdf': {'download': True, 'icon': '📕', 'name': 'PDF'},
    'application/vnd.openxmlformats-officedocument.wordprocessingml.document': {'download': True, 'icon': '📘', 'name': 'Word'},
    'application/msword': {'download': True, 'icon': '📘', 'name': 'Word'},
    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet': {'download': True, 'icon': '📗', 'name': 'Excel'},
    'text/plain': {'download': True, 'icon': '📝', 'name': 'Text'},
    'text/markdown': {'download': True, 'icon': '📝', 'name': 'Markdown'},
    'application/json': {'download': True, 'icon': '📝', 'name': 'JSON'},
}

def authenticate():
    """Authenticate with Google Drive API"""
    creds = None

    if os.path.exists(TOKEN_FILE):
        creds = Credentials.from_authorized_user_file(TOKEN_FILE, SCOPES)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            if not os.path.exists('credentials.json'):
                print("❌ ERROR: credentials.json not found!")
                print("\n📋 Setup Instructions:")
                print("1. Go to: https://console.cloud.google.com/apis/library/drive.googleapis.com")
                print("2. Click 'Enable'")
                print("3. Go to: https://console.cloud.google.com/apis/credentials")
                print("4. Click 'Create Credentials' → 'OAuth 2.0 Client ID'")
                print("5. Application type: 'Desktop app'")
                print("6. Download JSON and save as 'credentials.json' in this folder")
                sys.exit(1)

            flow = InstalledAppFlow.from_client_secrets_file('credentials.json', SCOPES)
            creds = flow.run_local_server(port=0)

        with open(TOKEN_FILE, 'w') as token:
            token.write(creds.to_json())

    return build('drive', 'v3', credentials=creds)


def generate_node_id() -> str:
    """Generate a unique node ID"""
    import uuid
    return f"n_{uuid.uuid4().hex[:8]}"


def chunk_text(text: str, chunk_size: int = DEFAULT_CHUNK_SIZE) -> List[Dict[str, Any]]:
    """
    Split text into chunks for RAG.
    Tries to split on paragraph boundaries.

    Returns list of chunks with metadata.
    """
    if not text or not text.strip():
        return []

    # Normalize whitespace
    text = re.sub(r'\n{3,}', '\n\n', text.strip())

    # If text is small enough, return as single chunk
    if len(text) <= chunk_size:
        return [{
            'text': text,
            'charCount': len(text),
            'isLeaf': True
        }]

    chunks = []
    paragraphs = text.split('\n\n')
    current_chunk = ""

    for para in paragraphs:
        para = para.strip()
        if not para:
            continue

        # If adding this paragraph would exceed chunk size
        if len(current_chunk) + len(para) + 2 > chunk_size:
            # Save current chunk if not empty
            if current_chunk:
                chunks.append({
                    'text': current_chunk.strip(),
                    'charCount': len(current_chunk.strip()),
                    'isLeaf': True
                })
                current_chunk = ""

            # If paragraph itself is too large, split it
            if len(para) > chunk_size:
                # Split on sentence boundaries
                sentences = re.split(r'(?<=[.!?])\s+', para)
                for sentence in sentences:
                    if len(current_chunk) + len(sentence) + 1 > chunk_size:
                        if current_chunk:
                            chunks.append({
                                'text': current_chunk.strip(),
                                'charCount': len(current_chunk.strip()),
                                'isLeaf': True
                            })
                            current_chunk = ""
                        # If single sentence is still too large, force split
                        if len(sentence) > chunk_size:
                            for i in range(0, len(sentence), chunk_size):
                                chunk_text = sentence[i:i+chunk_size].strip()
                                if chunk_text:
                                    chunks.append({
                                        'text': chunk_text,
                                        'charCount': len(chunk_text),
                                        'isLeaf': True
                                    })
                        else:
                            current_chunk = sentence
                    else:
                        current_chunk += " " + sentence if current_chunk else sentence
            else:
                current_chunk = para
        else:
            current_chunk += "\n\n" + para if current_chunk else para

    # Don't forget the last chunk
    if current_chunk.strip():
        chunks.append({
            'text': current_chunk.strip(),
            'charCount': len(current_chunk.strip()),
            'isLeaf': True
        })

    return chunks


def extract_text_from_pdf(content: bytes) -> str:
    """Extract text from PDF content"""
    if not HAS_PYMUPDF:
        return "[PDF extraction requires PyMuPDF: pip install PyMuPDF]"

    try:
        doc = fitz.open(stream=content, filetype="pdf")
        text_parts = []
        for page in doc:
            text_parts.append(page.get_text())
        doc.close()
        return "\n\n".join(text_parts)
    except Exception as e:
        return f"[PDF extraction error: {e}]"


def extract_text_from_docx(content: bytes) -> str:
    """Extract text from Word document"""
    if not HAS_DOCX:
        return "[Word extraction requires python-docx: pip install python-docx]"

    try:
        doc = DocxDocument(io.BytesIO(content))
        paragraphs = [para.text for para in doc.paragraphs if para.text.strip()]
        return "\n\n".join(paragraphs)
    except Exception as e:
        return f"[Word extraction error: {e}]"


def extract_text_from_xlsx(content: bytes) -> str:
    """Extract text from Excel spreadsheet"""
    if not HAS_OPENPYXL:
        return "[Excel extraction requires openpyxl: pip install openpyxl]"

    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        text_parts = []

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            text_parts.append(f"## Sheet: {sheet_name}\n")

            rows = []
            for row in sheet.iter_rows(values_only=True):
                cell_values = [str(cell) if cell is not None else "" for cell in row]
                if any(cell_values):
                    rows.append(" | ".join(cell_values))

            text_parts.append("\n".join(rows))

        wb.close()
        return "\n\n".join(text_parts)
    except Exception as e:
        return f"[Excel extraction error: {e}]"


def download_and_extract(service, file_id: str, mime_type: str, file_name: str) -> Optional[str]:
    """
    Download file content and extract text.

    For Google Workspace files: Use export API
    For binary files: Download and parse
    """
    type_info = EXTRACTABLE_TYPES.get(mime_type)
    if not type_info:
        return None

    try:
        # Google Workspace files - export as text
        if 'export' in type_info:
            export_mime = type_info['export']
            request = service.files().export_media(fileId=file_id, mimeType=export_mime)
            content = io.BytesIO()
            downloader = MediaIoBaseDownload(content, request)
            done = False
            while not done:
                _, done = downloader.next_chunk()
            return content.getvalue().decode('utf-8', errors='replace')

        # Binary files - download and parse
        elif 'download' in type_info:
            request = service.files().get_media(fileId=file_id)
            content = io.BytesIO()
            downloader = MediaIoBaseDownload(content, request)
            done = False
            while not done:
                _, done = downloader.next_chunk()

            raw_content = content.getvalue()

            # PDF
            if mime_type == 'application/pdf':
                return extract_text_from_pdf(raw_content)

            # Word
            elif mime_type in ['application/vnd.openxmlformats-officedocument.wordprocessingml.document',
                              'application/msword']:
                return extract_text_from_docx(raw_content)

            # Excel
            elif mime_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
                return extract_text_from_xlsx(raw_content)

            # Plain text files
            else:
                return raw_content.decode('utf-8', errors='replace')

    except Exception as e:
        print(f"    ⚠️  Extraction failed for {file_name}: {e}")
        return None

    return None


def create_chunk_nodes(chunks: List[Dict], parent_name: str, file_id: str) -> List[Dict]:
    """Create child nodes for each chunk"""
    nodes = []
    for i, chunk in enumerate(chunks):
        node = {
            'id': generate_node_id(),
            'name': f"Chunk {i + 1}" if len(chunks) > 1 else parent_name,
            'description': chunk['text'],
            'type': 'item',
            'icon': '📝',
            'external': {
                'type': 'gdrive:chunk',
                'fileId': file_id,
                'chunkIndex': i
            },
            '_rag': {
                'chunk': {
                    'charCount': chunk['charCount'],
                    'isLeaf': chunk['isLeaf']
                }
            }
        }
        nodes.append(node)
    return nodes


def scan_and_extract(service, folder_id: str = 'root', folder_name: str = 'My Drive',
                     depth: int = 0, max_depth: int = 5, chunk_size: int = DEFAULT_CHUNK_SIZE,
                     stats: Dict = None) -> List[Dict]:
    """
    Recursively scan folder and extract content from files.

    Returns list of nodes in knowledge-base pattern format.
    """
    if stats is None:
        stats = {'files_processed': 0, 'files_extracted': 0, 'total_chunks': 0, 'errors': 0}

    if depth > max_depth:
        print(f"{'  ' * depth}⚠️  Max depth {max_depth} reached")
        return []

    indent = '  ' * depth
    print(f"{indent}📂 {folder_name}")

    try:
        # Query files in this folder
        query = f"'{folder_id}' in parents and trashed=false"
        results = service.files().list(
            q=query,
            pageSize=100,  # Smaller batches for content extraction
            fields="files(id, name, mimeType, size, modifiedTime, webViewLink)"
        ).execute()

        items = results.get('files', [])
        children = []

        for item in items:
            is_folder = item['mimeType'] == 'application/vnd.google-apps.folder'
            type_info = EXTRACTABLE_TYPES.get(item['mimeType'], {})
            icon = type_info.get('icon', '📄' if not is_folder else '📁')

            if is_folder:
                # Recursively scan subfolders
                subchildren = scan_and_extract(
                    service, item['id'], item['name'],
                    depth + 1, max_depth, chunk_size, stats
                )

                if subchildren:  # Only add folders with extractable content
                    folder_node = {
                        'id': generate_node_id(),
                        'name': item['name'],
                        'type': 'phase',
                        'icon': '📁',
                        'external': {
                            'type': 'gdrive:folder',
                            'id': item['id']
                        },
                        'children': subchildren
                    }
                    children.append(folder_node)

            elif item['mimeType'] in EXTRACTABLE_TYPES:
                # Extract content from file
                stats['files_processed'] += 1
                print(f"{indent}  📄 {item['name'][:40]}...", end=" ")

                text = download_and_extract(service, item['id'], item['mimeType'], item['name'])

                if text and len(text.strip()) > 50:  # Skip near-empty files
                    stats['files_extracted'] += 1
                    chunks = chunk_text(text, chunk_size)
                    stats['total_chunks'] += len(chunks)

                    print(f"✓ {len(chunks)} chunks")

                    # Create file node with chunks as children
                    file_node = {
                        'id': generate_node_id(),
                        'name': item['name'],
                        'type': 'item',
                        'icon': icon,
                        'external': {
                            'type': 'gdrive:file',
                            'id': item['id']
                        },
                        '_rag': {
                            'source': {
                                'type': 'google-drive',
                                'fileId': item['id'],
                                'fileName': item['name'],
                                'mimeType': item['mimeType'],
                                'modifiedTime': item.get('modifiedTime', ''),
                                'extractedAt': datetime.now().isoformat()
                            }
                        },
                        'fileUrl': item.get('webViewLink', '')
                    }

                    # Add chunks as children if multiple
                    if len(chunks) > 1:
                        file_node['items'] = create_chunk_nodes(chunks, item['name'], item['id'])
                    elif len(chunks) == 1:
                        # Single chunk - embed in description
                        file_node['description'] = chunks[0]['text']
                        file_node['_rag']['chunk'] = chunks[0]

                    children.append(file_node)
                else:
                    print(f"⊘ (empty/small)")
            else:
                # Skip non-extractable files silently
                pass

        return children

    except Exception as e:
        print(f"{indent}❌ Error: {e}")
        stats['errors'] += 1
        return []


def export_gdrive_content(folder_id: str = 'root', max_depth: int = 5,
                          chunk_size: int = DEFAULT_CHUNK_SIZE) -> str:
    """Main export function"""
    print("\n🧠 TreeListy Google Drive Content Extractor")
    print("=" * 60)
    print(f"Folder: {folder_id}")
    print(f"Max depth: {max_depth} levels")
    print(f"Chunk size: {chunk_size} characters\n")

    # Check extraction libraries
    print("📚 Extraction Libraries:")
    print(f"   PyMuPDF (PDF): {'✅' if HAS_PYMUPDF else '❌'}")
    print(f"   python-docx (Word): {'✅' if HAS_DOCX else '❌'}")
    print(f"   openpyxl (Excel): {'✅' if HAS_OPENPYXL else '❌'}")
    print()

    # Authenticate
    print("🔐 Authenticating...")
    service = authenticate()
    print("✅ Authenticated\n")

    # Get folder name if not root
    folder_name = "My Drive"
    if folder_id != 'root':
        try:
            folder_meta = service.files().get(fileId=folder_id, fields="name").execute()
            folder_name = folder_meta.get('name', folder_id)
        except:
            pass

    # Scan and extract
    print("📥 Scanning and extracting content...\n")
    stats = {'files_processed': 0, 'files_extracted': 0, 'total_chunks': 0, 'errors': 0}
    children = scan_and_extract(service, folder_id, folder_name, 0, max_depth, chunk_size, stats)

    # Create TreeListy structure (knowledge-base pattern)
    tree = {
        'id': 'kb-gdrive',
        'name': f'📚 {folder_name} Knowledge Base',
        'type': 'root',
        'icon': '📚',
        'expanded': True,
        'pattern': {
            'key': 'knowledge-base'
        },
        'source': {
            'type': 'google-drive',
            'folderId': folder_id,
            'folderName': folder_name,
            'lastSync': datetime.now().isoformat(),
            'syncDepth': max_depth,
            'chunkSize': chunk_size
        },
        '_rag': {
            'stats': {
                'filesProcessed': stats['files_processed'],
                'filesExtracted': stats['files_extracted'],
                'totalChunks': stats['total_chunks'],
                'errors': stats['errors']
            }
        },
        'children': children
    }

    # Save to file
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_file = f'gdrive-content-{timestamp}.json'

    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(tree, f, indent=2, ensure_ascii=False)

    # Summary
    print("\n" + "=" * 60)
    print(f"✅ SUCCESS! Exported to: {output_file}")
    print(f"\n📊 Statistics:")
    print(f"   Files processed: {stats['files_processed']}")
    print(f"   Files extracted: {stats['files_extracted']}")
    print(f"   Total chunks: {stats['total_chunks']}")
    print(f"   Errors: {stats['errors']}")
    print(f"\n📋 Next Steps:")
    print(f"   1. Open TreeListy in browser")
    print(f"   2. Click '📂 Import' → Select '{output_file}'")
    print(f"   3. Pattern auto-detected: '📚 Knowledge Base'")
    print(f"   4. Use search or TreeBeard to query your documents!")
    print("=" * 60)

    return output_file


if __name__ == '__main__':
    parser = argparse.ArgumentParser(
        description='Extract text content from Google Drive files for RAG'
    )
    parser.add_argument(
        'folder_id',
        nargs='?',
        default='root',
        help='Google Drive folder ID (default: root = My Drive)'
    )
    parser.add_argument(
        '--max-depth', '-d',
        type=int,
        default=5,
        help='Maximum folder depth to scan (default: 5)'
    )
    parser.add_argument(
        '--chunk-size', '-c',
        type=int,
        default=DEFAULT_CHUNK_SIZE,
        help=f'Target chunk size in characters (default: {DEFAULT_CHUNK_SIZE})'
    )

    args = parser.parse_args()

    # Validate chunk size
    chunk_size = max(MIN_CHUNK_SIZE, min(MAX_CHUNK_SIZE, args.chunk_size))
    if chunk_size != args.chunk_size:
        print(f"⚠️  Chunk size adjusted to {chunk_size} (range: {MIN_CHUNK_SIZE}-{MAX_CHUNK_SIZE})")

    export_gdrive_content(args.folder_id, args.max_depth, chunk_size)
